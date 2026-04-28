"""
Lambda 2 — facturacion-orquestador
Trigger: S3 Event en processed/ (Excel extraido por Lambda 1)
Flujo: Lee Excel -> agrupa por Nro documento -> Bedrock genera CxC -> PDF -> S3 output/
Env vars: BUCKET_NAME, TABLE_NAME, MODEL_ID, PDF_BUCKET, DESTINATARIO
Layer: hacksiesa-deps-full (xlrd + openpyxl + fpdf2)
"""
import os, io, json, uuid, boto3, re
from datetime import datetime
from urllib.parse import unquote_plus
from fpdf import FPDF

BUCKET = os.environ['BUCKET_NAME']
PDF_BUCKET = os.environ.get('PDF_BUCKET', 'financial-723595585512')
TABLE = os.environ['TABLE_NAME']
MODEL_ID = os.environ.get('MODEL_ID', 'us.anthropic.claude-3-haiku-20240307-v1:0')
DESTINATARIO = os.environ.get('DESTINATARIO', 'juan.daza@siesa.com')

s3 = boto3.client('s3')
dynamodb = boto3.resource('dynamodb')
bedrock = boto3.client('bedrock-runtime', region_name='us-east-1')

SYSTEM_PROMPT = """Eres un Agente de IA experto en facturacion colombiana.

TAREA: Genera una cuenta de cobro para UN documento especifico con sus items.

REGLAS:
1. NUNCA inventes datos.
2. "Desc. sucursal factura" = nombre del cliente.
3. "Nro documento" = numero del pedido/factura.
4. "Desc. corta item" = descripcion del servicio/producto.
5. "Total" = valor de cada item.
6. "Notas docto" = observaciones (periodo, productos, etc).
7. valor_total = SUMA de todos los items del documento.
8. El email_body debe ser profesional, en espanol, mencionando los documentos.

DATOS DEL EMISOR:
- Nombre: SIESA HACK S.A.S
- NIT: 900.123.456-7
- Direccion: Calle 10 #5-30, Cali
- Telefono: (602) 555-0100

RESPONDE SOLO JSON VALIDO (sin markdown):
{
  "status": "success",
  "pdf_content": {
    "titulo": "Cuenta de cobro",
    "numero": "CC-2026-XXXX",
    "fecha_emision": "YYYY-MM-DD",
    "emisor": {"nombre": "SIESA HACK S.A.S", "nit": "900.123.456-7", "direccion": "Calle 10 #5-30, Cali", "telefono": "(602) 555-0100"},
    "cliente": {"nombre": "[del Excel]", "identificacion": "N/A", "direccion": "N/A"},
    "nro_documento": "[Nro documento del Excel]",
    "items": [{"descripcion": "[Desc. corta item]", "valor": 0.00}],
    "valor_total": 0.00,
    "valor_en_letras": "[monto en palabras en espanol]",
    "observaciones": "[De Notas docto]"
  },
  "email_body": "<html><body>correo profesional</body></html>",
  "scheduled_date": "YYYY-MM-DDTHH:MM:SSZ",
  "error_report": null
}

REGLAS PARA scheduled_date: 3 dias habiles desde hoy, excluir fines de semana, hora 08:00 UTC."""


class CuentaCobroPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 18)
        self.cell(0, 12, "CUENTA DE COBRO", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(4); self.set_draw_color(30, 80, 120); self.set_line_width(0.5)
        self.line(10, self.get_y(), 200, self.get_y()); self.ln(8)
    def footer(self):
        self.set_y(-15); self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}", align="C")

def generar_pdf(pc):
    pdf = CuentaCobroPDF(); pdf.add_page(); pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(95, 8, f'No: {pc.get("numero", "N/A")}')
    pdf.cell(95, 8, f'Fecha: {pc.get("fecha_emision", "N/A")}', align="R", new_x="LMARGIN", new_y="NEXT")
    if pc.get("nro_documento"):
        pdf.cell(0, 8, f'Documento: {pc["nro_documento"]}', new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)
    e = pc.get("emisor", {}); pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "DATOS DEL EMISOR", new_x="LMARGIN", new_y="NEXT"); pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, e.get("nombre", ""), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f'NIT: {e.get("nit", "")}', new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f'{e.get("direccion", "")} | Tel: {e.get("telefono", "")}', new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)
    c = pc.get("cliente", {}); pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "DATOS DEL CLIENTE", new_x="LMARGIN", new_y="NEXT"); pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, c.get("nombre", "N/A"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f'ID: {c.get("identificacion", "N/A")}', new_x="LMARGIN", new_y="NEXT"); pdf.ln(6)
    pdf.set_font("Helvetica", "B", 10); pdf.set_fill_color(30, 80, 120); pdf.set_text_color(255, 255, 255)
    pdf.cell(130, 8, "Descripcion", border=1, fill=True)
    pdf.cell(60, 8, "Valor", border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 10)
    for item in pc.get("items", []):
        pdf.cell(130, 8, str(item.get("descripcion", ""))[:60], border=1)
        pdf.cell(60, 8, f"${item.get('valor', 0):,.2f}", border=1, align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(130, 10, "TOTAL", border=1)
    pdf.cell(60, 10, f'${pc.get("valor_total", 0):,.2f}', border=1, align="R", new_x="LMARGIN", new_y="NEXT")
    if pc.get("valor_en_letras"):
        pdf.ln(4); pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 6, f'Son: {pc["valor_en_letras"]}', new_x="LMARGIN", new_y="NEXT")
    if pc.get("observaciones"):
        pdf.ln(4); pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(0, 5, f'Obs: {str(pc["observaciones"])[:200]}')
    return pdf.output()


def leer_excel_s3(s3_key):
    obj = s3.get_object(Bucket=BUCKET, Key=s3_key); data = obj['Body'].read()
    if s3_key.lower().endswith('.xls'):
        import xlrd, tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xls', delete=False); tmp.write(data); tmp.close()
        wb = xlrd.open_workbook(tmp.name); ws = wb.sheet_by_index(0)
        header_row = 0
        for r in range(min(10, ws.nrows)):
            if sum(1 for c in range(ws.ncols) if ws.cell_value(r, c)) >= 3: header_row = r; break
        headers = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
        print(f"  Headers fila {header_row}: {headers}")
        rows, last_c, last_d, last_n = [], "", "", ""
        for r in range(header_row + 1, ws.nrows):
            row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
            if row.get(headers[0]) and str(row[headers[0]]).strip(): last_c = str(row[headers[0]]).strip()
            else: row[headers[0]] = last_c
            if 'Nro documento' in row:
                if row['Nro documento'] and str(row['Nro documento']).strip(): last_d = str(row['Nro documento']).strip()
                else: row['Nro documento'] = last_d
            if 'Notas docto' in row:
                if row['Notas docto'] and str(row['Notas docto']).strip(): last_n = str(row['Notas docto']).strip()
                else: row['Notas docto'] = last_n
            if 'total' in str(row.get(headers[0], '')).lower(): continue
            rows.append(row)
        return headers, rows
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True); ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True)); header_row = 0
        for i, row in enumerate(all_rows[:10]):
            if sum(1 for v in row if v) >= 3: header_row = i; break
        headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(all_rows[header_row])]
        rows, last_c, last_d, last_n = [], "", "", ""
        for row in all_rows[header_row+1:]:
            d = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
            if d.get(headers[0]) and d[headers[0]].strip(): last_c = d[headers[0]].strip()
            else: d[headers[0]] = last_c
            if 'Nro documento' in d:
                if d['Nro documento'].strip(): last_d = d['Nro documento'].strip()
                else: d['Nro documento'] = last_d
            if 'Notas docto' in d:
                if d['Notas docto'].strip(): last_n = d['Notas docto'].strip()
                else: d['Notas docto'] = last_n
            if 'total' in d.get(headers[0], '').lower(): continue
            rows.append(d)
        return headers, rows

def agrupar_por_documento(rows):
    grupos = {}
    for row in rows:
        nro = str(row.get('Nro documento', '')).strip()
        if not nro: continue
        if nro not in grupos:
            grupos[nro] = {'cliente': str(row.get('Desc. sucursal factura', row.get(list(row.keys())[0], '?'))).strip(),
                'nro_documento': nro, 'notas': str(row.get('Notas docto', '')).strip(), 'items': []}
        grupos[nro]['items'].append(row)
    return grupos

def llamar_bedrock(doc_data):
    try:
        user_msg = f"Fecha actual: {datetime.now().strftime('%Y-%m-%d')}\nCliente: {doc_data['cliente']}\nNro documento: {doc_data['nro_documento']}\nNotas: {doc_data['notas']}\nItems ({len(doc_data['items'])}):\n{json.dumps(doc_data['items'], default=str, ensure_ascii=False, indent=2)}\n\nGenera la cuenta de cobro."
        response = bedrock.invoke_model(modelId=MODEL_ID, contentType='application/json',
            body=json.dumps({'anthropic_version': 'bedrock-2023-05-31', 'max_tokens': 4096,
                'system': SYSTEM_PROMPT, 'messages': [{'role': 'user', 'content': user_msg}]}))
        text = json.loads(response['body'].read())['content'][0]['text']
        text = re.sub(r'^```json\s*', '', text.strip()); text = re.sub(r'\s*```$', '', text.strip())
        return json.loads(text)
    except Exception as e:
        print(f"  Error Bedrock: {e}")
        return {'status': 'error', 'error_report': {'mensaje': str(e)}, 'pdf_content': None, 'email_body': None, 'scheduled_date': None}


def lambda_handler(event, context):
    record = event['Records'][0]; s3_key = unquote_plus(record['s3']['object']['key'])
    print(f"Key: {s3_key}")
    if not s3_key.lower().endswith(('.xlsx', '.xls', '.xlsm')): return {'statusCode': 200}

    remitente = 'desconocido'
    try:
        meta = s3.head_object(Bucket=BUCKET, Key=s3_key)
        remitente = meta.get('Metadata', {}).get('remitente', 'desconocido')
    except: pass
    print(f"Remitente: {remitente} | Model: {MODEL_ID}")

    try:
        headers, rows = leer_excel_s3(s3_key); print(f"Excel: {len(rows)} filas")
    except Exception as e:
        print(f"Error Excel: {e}"); return {'statusCode': 500}

    grupos = agrupar_por_documento(rows); print(f"Documentos: {len(grupos)}")
    if not grupos: return {'statusCode': 200}

    table = dynamodb.Table(TABLE); resultados = []
    for nro_doc, doc_data in grupos.items():
        print(f"\n  Doc: {nro_doc} | {doc_data['cliente']} | {len(doc_data['items'])} items")
        agent_output = llamar_bedrock(doc_data); print(f"  Bedrock: {agent_output.get('status')}")
        factura_id = str(uuid.uuid4())
        json_key = f'output/{factura_id}.json'
        full_data = {'factura_id': factura_id, 'nro_documento': nro_doc, 'cliente_excel': doc_data['cliente'],
            'items_count': len(doc_data['items']), 'archivo_origen': s3_key, 'remitente': remitente,
            'destinatario': DESTINATARIO, **agent_output}
        s3.put_object(Bucket=BUCKET, Key=json_key, Body=json.dumps(full_data, ensure_ascii=False, indent=2), ContentType='application/json')
        if agent_output.get('status') == 'success' and agent_output.get('pdf_content'):
            try:
                pdf_bytes = generar_pdf(agent_output['pdf_content']); pdf_key = f'output/{factura_id}.pdf'
                s3.put_object(Bucket=PDF_BUCKET, Key=pdf_key, Body=pdf_bytes, ContentType='application/pdf')
                print(f"  PDF: {pdf_key} ({len(pdf_bytes):,} bytes)")
            except Exception as e: print(f"  Error PDF: {e}")
        table.put_item(Item={'factura_id': factura_id, 'nro_documento': nro_doc, 'cliente': doc_data['cliente'],
            'estado': agent_output.get('status', 'error'), 'fecha_procesamiento': datetime.now().isoformat(),
            'fecha_envio_programado': agent_output.get('scheduled_date', '') or '',
            's3_excel_key': s3_key, 's3_output_key': json_key, 'remitente': remitente, 'destinatario': DESTINATARIO})
        resultados.append({'factura_id': factura_id, 'nro_documento': nro_doc, 'status': agent_output.get('status')})

    print(f"\nTotal: {len(resultados)}")
    return {'statusCode': 200, 'resultados': resultados}
